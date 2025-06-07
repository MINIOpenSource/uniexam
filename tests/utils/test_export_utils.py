# -*- coding: utf-8 -*-
"""
app.utils.export_utils 模块的单元测试。
(Unit tests for the app.utils.export_utils module.)
"""

import csv
import io

import openpyxl  # 用于解析XLSX内容 (For parsing XLSX content)
import pytest
from fastapi.responses import StreamingResponse

from app.utils.export_utils import data_to_csv, data_to_xlsx

# region 辅助函数 (Helper Functions)


async def _read_streaming_response_content(response: StreamingResponse) -> bytes:
    """
    异步读取 StreamingResponse 的完整内容。
    (Asynchronously reads the full content of a StreamingResponse.)
    """
    content_bytes = b""
    async for chunk in response.body_iterator:
        content_bytes += chunk
    return content_bytes


# endregion

# region data_to_csv 测试 (data_to_csv Tests)


@pytest.mark.asyncio
async def test_data_to_csv_empty_data():
    """测试 data_to_csv 处理空数据列表的情况。"""
    headers = ["栏位一", "栏位二", "栏位三"]  # (Column1, Column2, Column3)
    filename = "empty_export.csv"

    response = data_to_csv(data_list=[], headers=headers, filename=filename)

    assert isinstance(response, StreamingResponse), "返回类型不正确。"
    assert response.media_type == "text/csv", "媒体类型不正确。"
    assert (
        response.headers["Content-Disposition"] == f'attachment; filename="{filename}"'
    ), "Content-Disposition 头部不正确。"

    content_bytes = await _read_streaming_response_content(response)
    # utf-8-sig 包含BOM: \xef\xbb\xbf
    # (utf-8-sig includes BOM: \xef\xbb\xbf)
    assert content_bytes.startswith(b"\xef\xbb\xbf"), "CSV内容应以UTF-8 BOM开头。"

    # 解码并解析CSV内容 (Decode and parse CSV content)
    csv_content_str = content_bytes.decode("utf-8-sig")
    csv_reader = csv.reader(io.StringIO(csv_content_str))

    parsed_rows = list(csv_reader)
    assert len(parsed_rows) == 1, "空数据CSV应只包含一行表头。"
    assert parsed_rows[0] == headers, "CSV表头与预期不符。"


@pytest.mark.asyncio
async def test_data_to_csv_with_data():
    """测试 data_to_csv 处理包含数据的列表。"""
    headers = ["名称", "值", "描述"]  # (Name, Value, Description)
    data = [
        {"名称": "项目A", "值": 100, "描述": "这是项目A的描述"},
        {"名称": "项目B", "值": 250, "描述": "项目B有一些特性"},
    ]
    filename = "data_export.csv"

    response = data_to_csv(data_list=data, headers=headers, filename=filename)
    content_bytes = await _read_streaming_response_content(response)
    csv_content_str = content_bytes.decode("utf-8-sig")
    csv_reader = csv.reader(io.StringIO(csv_content_str))

    parsed_rows = list(csv_reader)
    assert len(parsed_rows) == 3, "CSV行数不正确 (应为1行表头 + 2行数据)。"
    assert parsed_rows[0] == headers, "CSV表头不正确。"

    # 验证数据行 (Verify data rows)
    assert parsed_rows[1] == ["项目A", "100", "这是项目A的描述"], "第一行数据不正确。"
    assert parsed_rows[2] == ["项目B", "250", "项目B有一些特性"], "第二行数据不正确。"
    # 注意：CSV中的所有值默认都会被读取为字符串 (Note: All values in CSV are read as strings by default)


@pytest.mark.asyncio
async def test_data_to_csv_special_characters():
    """测试 data_to_csv 处理包含特殊字符（逗号, 引号, 换行, 中文）的数据。"""
    headers = ["ID", "文本内容"]  # (ID, Text Content)
    data = [
        {"ID": 1, "文本内容": '包含逗号,和"引号"的文本。'},
        {"ID": 2, "文本内容": "跨多行的\n文本内容，\r\n以及中文句子。"},
        {"ID": 3, "文本内容": None},  # 测试None值 (Test None value)
    ]
    filename = "special_chars.csv"

    response = data_to_csv(data_list=data, headers=headers, filename=filename)
    content_bytes = await _read_streaming_response_content(response)
    csv_content_str = content_bytes.decode("utf-8-sig")

    # 直接比较生成的CSV字符串的期望值，因为csv.reader的解析可能隐藏一些细节
    # (Directly compare expected value of generated CSV string, as csv.reader parsing might hide details)
    # 期望的CSV输出 (Expected CSV output)
    # ID,文本内容
    # 1,"包含逗号,和""引号""的文本。"  <-- 注意双引号的转义 (Note escaping of double quotes)
    # 2,"跨多行的
    # 文本内容，
    # 以及中文句子。"
    # 3,"" <-- None 值应为空字符串 (None value should be empty string)

    # 使用 io.StringIO 和 csv.writer 来构建期望的输出，以确保与实现逻辑一致
    # (Use io.StringIO and csv.writer to build expected output for consistency with implementation logic)
    expected_output_io = io.StringIO()
    csv_writer = csv.writer(expected_output_io)
    csv_writer.writerow(headers)
    csv_writer.writerow([data[0]["ID"], data[0]["文本内容"]])
    csv_writer.writerow([data[1]["ID"], data[1]["文本内容"]])
    csv_writer.writerow([data[2]["ID"], ""])  # None 应该被转换为空字符串
    # (None should be converted to empty string)
    expected_csv_str = expected_output_io.getvalue()

    assert csv_content_str == expected_csv_str, "包含特殊字符的CSV内容与预期不符。"


# endregion

# region data_to_xlsx 测试 (data_to_xlsx Tests)


@pytest.mark.asyncio
async def test_data_to_xlsx_empty_data():
    """测试 data_to_xlsx 处理空数据列表的情况。"""
    headers = ["ID", "名称", "数量"]  # (ID, Name, Quantity)
    filename = "empty_export.xlsx"

    response = data_to_xlsx(data_list=[], headers=headers, filename=filename)

    assert isinstance(response, StreamingResponse), "返回类型不正确。"
    assert (
        response.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ), "媒体类型不正确。"
    assert (
        response.headers["Content-Disposition"] == f'attachment; filename="{filename}"'
    ), "Content-Disposition 头部不正确。"

    content_bytes = await _read_streaming_response_content(response)

    # 解析XLSX内容 (Parse XLSX content)
    workbook = openpyxl.load_workbook(io.BytesIO(content_bytes))
    sheet = workbook.active

    assert sheet.max_row == 1, "空数据XLSX应只包含一行表头。"
    header_row_values = [cell.value for cell in sheet[1]]
    assert header_row_values == headers, "XLSX表头与预期不符。"


@pytest.mark.asyncio
async def test_data_to_xlsx_with_data():
    """测试 data_to_xlsx 处理包含数据的列表。"""
    headers = ["产品名称", "单价", "库存"]  # (Product Name, Unit Price, Stock)
    data = [
        {"产品名称": "笔记本电脑", "单价": 7500, "库存": 50},
        {"产品名称": "智能手机", "单价": 4200, "库存": 120},
        {
            "产品名称": "蓝牙耳机",
            "单价": None,
            "库存": 300,
        },  # 测试None值 (Test None value)
    ]
    filename = "data_export.xlsx"

    response = data_to_xlsx(data_list=data, headers=headers, filename=filename)
    content_bytes = await _read_streaming_response_content(response)
    workbook = openpyxl.load_workbook(io.BytesIO(content_bytes))
    sheet = workbook.active

    assert sheet.max_row == 4, "XLSX行数不正确 (1表头 + 3数据)。"

    # 验证表头 (Verify headers)
    header_row_values_sheet = [cell.value for cell in sheet[1]]
    assert header_row_values_sheet == headers, "XLSX表头不正确。"

    # 验证数据行 (Verify data rows)
    # 第1行数据 (Row 1 data)
    row2_values = [cell.value for cell in sheet[2]]
    assert row2_values == ["笔记本电脑", 7500, 50], "XLSX第一行数据不正确。"

    # 第2行数据 (Row 2 data)
    row3_values = [cell.value for cell in sheet[3]]
    assert row3_values == ["智能手机", 4200, 120], "XLSX第二行数据不正确。"

    # 第3行数据 (含None) (Row 3 data (with None))
    row4_values = [cell.value for cell in sheet[4]]
    assert row4_values == ["蓝牙耳机", None, 300], "XLSX第三行数据（含None）不正确。"


@pytest.mark.asyncio
async def test_data_to_xlsx_data_types():
    """测试 data_to_xlsx 处理不同基本数据类型。"""
    headers = ["字符串", "整数", "浮点数", "布尔值", "空值"]
    # (String, Integer, Float, Boolean, Null Value)
    data = [
        {
            "字符串": "你好世界",
            "整数": 123,
            "浮点数": 45.67,
            "布尔值": True,
            "空值": None,
        },
        {"字符串": "Hello", "整数": -5, "浮点数": 0.001, "布尔值": False, "空值": None},
    ]
    filename = "datatypes.xlsx"

    response = data_to_xlsx(data_list=data, headers=headers, filename=filename)
    content_bytes = await _read_streaming_response_content(response)
    workbook = openpyxl.load_workbook(io.BytesIO(content_bytes))
    sheet = workbook.active

    # 验证第一行数据的数据类型和值
    # (Verify data types and values of the first data row)
    # openpyxl 会尝试保留原始Python类型，例如数字类型
    # (openpyxl will try to preserve original Python types, e.g., numeric types)
    row2_values_sheet = [cell.value for cell in sheet[2]]
    assert row2_values_sheet[0] == "你好世界"
    assert isinstance(row2_values_sheet[0], str), "字符串类型不正确。"
    assert row2_values_sheet[1] == 123
    assert isinstance(row2_values_sheet[1], int), "整数类型不正确。"
    assert row2_values_sheet[2] == 45.67
    assert isinstance(row2_values_sheet[2], float), "浮点数类型不正确。"
    assert row2_values_sheet[3] is True, (
        "布尔值True不正确。"
    )  # openpyxl stores bools correctly
    assert row2_values_sheet[4] is None, "空值不正确 (应为None)。"

    row3_values_sheet = [cell.value for cell in sheet[3]]
    assert row3_values_sheet[3] is False, "布尔值False不正确。"


# endregion
